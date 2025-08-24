#!/usr/bin/env python3
"""
Telegram File Processor - App esterna per processing PDF e Excel
Riceve webhook da Telegram, processa file complessi, inoltra dati a N8N
"""

import os
import io
import logging
import requests
from flask import Flask, request, jsonify
from werkzeug.exceptions import BadRequest
import PyPDF2
import pdfplumber
from openpyxl import load_workbook
import traceback
from datetime import datetime
import json
import re

# Configurazione logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configurazione
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
N8N_WEBHOOK_URL = os.getenv('N8N_WEBHOOK_URL')
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB limit
TELEGRAM_API_BASE = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"

def download_telegram_file(file_id):
    """Scarica un file da Telegram usando file_id"""
    try:
        # Ottieni info del file
        file_info_url = f"{TELEGRAM_API_BASE}/getFile"
        response = requests.get(file_info_url, params={'file_id': file_id})
        response.raise_for_status()
        
        file_info = response.json()
        if not file_info.get('ok'):
            raise Exception(f"Errore API Telegram: {file_info.get('description')}")
        
        file_path = file_info['result']['file_path']
        file_size = file_info['result'].get('file_size', 0)
        
        if file_size > MAX_FILE_SIZE:
            raise Exception(f"File troppo grande: {file_size} bytes")
        
        # Scarica il file
        download_url = f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}"
        file_response = requests.get(download_url)
        file_response.raise_for_status()
        
        return file_response.content, file_path
        
    except Exception as e:
        logger.error(f"Errore download file Telegram: {e}")
        raise

def process_excel_file(file_content, filename):
    """Processa file Excel ed estrae dati in formato tabellare"""
    try:
        # Carica il workbook direttamente dal contenuto
        workbook = load_workbook(io.BytesIO(file_content), read_only=True)
        
        # Prende il primo sheet
        sheet = workbook.active
        sheet_name = sheet.title
        
        # Estrae tutte le righe con dati
        data = []
        headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # Prima riga come headers
                headers = [str(cell) if cell is not None else f"Column_{i+1}" 
                          for i, cell in enumerate(row)]
            else:
                # Converte la riga in dizionario
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Converte tipi di dati
                        if cell is None:
                            value = ""
                        elif isinstance(cell, datetime):
                            value = cell.isoformat()
                        else:
                            value = str(cell)
                        row_data[headers[i]] = value
                
                # Aggiunge solo righe non vuote
                if any(row_data.values()):
                    data.append(row_data)
        
        result = {
            'type': 'excel',
            'filename': filename,
            'sheet_name': sheet_name,
            'headers': headers,
            'data': data,
            'row_count': len(data),
            'processed_at': datetime.now().isoformat()
        }
        
        logger.info(f"Excel processato: {len(data)} righe, {len(headers)} colonne")
        return result
        
    except Exception as e:
        logger.error(f"Errore processing Excel: {e}")
        raise

def process_pdf_file(file_content, filename):
    """Processa file PDF ed estrae testo e tabelle"""
    try:
        tables_data = []
        text_content = ""
        
        # Prova prima con pdfplumber (migliore per tabelle)
        try:
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # Estrae testo
                    page_text = page.extract_text()
                    if page_text:
                        text_content += f"\n--- Pagina {page_num + 1} ---\n{page_text}\n"
                    
                    # Estrae tabelle
                    tables = page.extract_tables()
                    for table_idx, table in enumerate(tables):
                        if table and len(table) > 1:  # Almeno header + 1 riga
                            headers = table[0] if table[0] else [f"Col_{i+1}" for i in range(len(table[1]))]
                            
                            table_data = []
                            for row in table[1:]:
                                if row and any(cell for cell in row):  # Riga non vuota
                                    row_dict = {}
                                    for i, cell in enumerate(row):
                                        if i < len(headers):
                                            row_dict[str(headers[i]) if headers[i] else f"Col_{i+1}"] = str(cell) if cell else ""
                                    table_data.append(row_dict)
                            
                            if table_data:
                                tables_data.append({
                                    'page': page_num + 1,
                                    'table': table_idx + 1,
                                    'headers': headers,
                                    'data': table_data
                                })
        
        except Exception as e:
            logger.warning(f"pdfplumber fallito, provo PyPDF2: {e}")
            
            # Fallback con PyPDF2 solo per testo
            try:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text_content += f"\n--- Pagina {page_num + 1} ---\n{page_text}\n"
            except Exception as e2:
                logger.error(f"Anche PyPDF2 fallito: {e2}")
                text_content = "Errore estrazione testo PDF"
        
        # Prova a trovare pattern tabellari nel testo se non ci sono tabelle estratte
        if not tables_data and text_content:
            potential_tables = extract_tables_from_text(text_content)
            tables_data.extend(potential_tables)
        
        result = {
            'type': 'pdf',
            'filename': filename,
            'text_content': text_content.strip(),
            'tables': tables_data,
            'tables_count': len(tables_data),
            'processed_at': datetime.now().isoformat()
        }
        
        logger.info(f"PDF processato: {len(tables_data)} tabelle, {len(text_content)} caratteri di testo")
        return result
        
    except Exception as e:
        logger.error(f"Errore processing PDF: {e}")
        raise

def extract_tables_from_text(text):
    """Cerca pattern tabellari nel testo estratto"""
    tables = []
    
    # Pattern per righe che sembrano tabelle (separatori comuni)
    lines = text.split('\n')
    potential_table_lines = []
    
    for line in lines:
        # Cerca righe con separatori multipli
        if len(re.findall(r'[\t|,;]', line)) >= 2:
            potential_table_lines.append(line.strip())
        elif len(line.split()) >= 3 and not line.strip().endswith('.'):
            # Righe con spazi che potrebbero essere colonne
            potential_table_lines.append(line.strip())
    
    # Se trova almeno 3 righe consecutive che sembrano tabellari
    if len(potential_table_lines) >= 3:
        # Determina il separatore più comune
        separators = ['\t', '|', ',', ';']
        best_sep = None
        max_cols = 0
        
        for sep in separators:
            avg_cols = sum(len(line.split(sep)) for line in potential_table_lines[:5]) / min(5, len(potential_table_lines))
            if avg_cols > max_cols and avg_cols >= 2:
                max_cols = avg_cols
                best_sep = sep
        
        if best_sep:
            headers = potential_table_lines[0].split(best_sep)
            table_data = []
            
            for line in potential_table_lines[1:]:
                cells = line.split(best_sep)
                if len(cells) >= len(headers):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        row_dict[header.strip()] = cells[i].strip() if i < len(cells) else ""
                    table_data.append(row_dict)
            
            if table_data:
                tables.append({
                    'page': 'text_extraction',
                    'table': 1,
                    'headers': [h.strip() for h in headers],
                    'data': table_data
                })
    
    return tables

def send_to_n8n(processed_data):
    """Invia i dati processati a N8N via webhook"""
    try:
        if not N8N_WEBHOOK_URL:
            raise Exception("N8N_WEBHOOK_URL non configurato")
        
        response = requests.post(
            N8N_WEBHOOK_URL,
            json=processed_data,
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        response.raise_for_status()
        
        logger.info(f"Dati inviati a N8N con successo: {response.status_code}")
        return True
        
    except Exception as e:
        logger.error(f"Errore invio a N8N: {e}")
        raise

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'telegram_token_configured': bool(TELEGRAM_BOT_TOKEN),
        'n8n_webhook_configured': bool(N8N_WEBHOOK_URL)
    })

@app.route('/webhook', methods=['POST'])
@app.route('/webhook', methods=['POST'])
def telegram_webhook():
    """Riceve dati da N8N con file già scaricato"""
    try:
        data = request.get_json()
        if not data:
            raise BadRequest("No JSON data received")
        
        # N8N invia i dati del file, non serve scaricare da Telegram
        file_content = data.get('file_content')  # Base64 del file
        filename = data.get('filename', 'unknown')
        file_type = data.get('file_type', 'unknown')
        
        if not file_content:
            return jsonify({'status': 'error', 'error': 'No file content provided'})
        
        # Processa il file
        if file_type in ['xlsx', 'xls']:
            processed_data = process_excel_file(base64.b64decode(file_content), filename)
        elif file_type == 'pdf':
            processed_data = process_pdf_file(base64.b64decode(file_content), filename)
        else:
            return jsonify({'status': 'error', 'error': f'Unsupported file type: {file_type}'})
        
        # Invia a N8N
        send_to_n8n(processed_data)
        
        return jsonify({'status': 'processed'})
        
    except Exception as e:
        logger.error(f"Errore webhook handler: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500
@app.route('/process-file', methods=['POST'])
def process_file_endpoint():
    """Endpoint alternativo per processing diretto tramite file_id"""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        file_type = data.get('file_type', 'auto')
        
        if not file_id:
            raise BadRequest("file_id richiesto")
        
        # Scarica il file
        file_content, file_path = download_telegram_file(file_id)
        filename = file_path.split('/')[-1]
        
        # Determina il tipo se non specificato
        if file_type == 'auto':
            if filename.lower().endswith(('.xlsx', '.xls')):
                file_type = 'excel'
            elif filename.lower().endswith('.pdf'):
                file_type = 'pdf'
            else:
                raise Exception(f"Tipo file non riconosciuto: {filename}")
        
        # Processa
        if file_type == 'excel':
            processed_data = process_excel_file(file_content, filename)
        elif file_type == 'pdf':
            processed_data = process_pdf_file(file_content, filename)
        else:
            raise Exception(f"Tipo file non supportato: {file_type}")
        
        # Invia a N8N
        send_to_n8n(processed_data)
        
        return jsonify({
            'status': 'processed',
            'data': processed_data
        })
        
    except Exception as e:
        logger.error(f"Errore process-file: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    logger.info(f"Avvio app su porta {port}")
    logger.info(f"Telegram token configurato: {bool(TELEGRAM_BOT_TOKEN)}")
    logger.info(f"N8N webhook configurato: {bool(N8N_WEBHOOK_URL)}")
    app.run(host='0.0.0.0', port=port, debug=False)
